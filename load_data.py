from time import sleep
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import date, datetime
import parse
import config

with open("codes.txt", 'r') as file:
    text = file.read().replace("'", '')
    file.close()

codes = [m.strip() for m in text.split(',')]

driver = parse.init_driver()

driver.get(url=config.url)

wait = parse.init_wait(driver)

if __name__ == "__main__":
    #  Авторизация
    is_auth = False
    try:
        is_auth = parse.auth(driver, wait, config.mail, config.password)
    except Exception as e:
        print(e)
        sleep(3)
        driver.get(url=config.url)
        is_auth = parse.auth(driver, wait, config.mail, config.password)

    #  Инициализация эксельки
    dest_filename = str(datetime.now().strftime("%Y-%m-%d_%H-%M-%S")) + ".xlsx"
    temp_wb = Workbook()
    temp_sheet = temp_wb.active
    workbook = load_workbook(filename="database.xlsx")
    sheet = workbook["test"]

    waste_codes = open("waste_codes.txt", 'w')

    # Считывание заголовков
    headers = {}
    column_count = 1
    while True:
        letter = get_column_letter(column_count)
        if sheet[f'{letter}1'].value:
            headers.update({sheet[f'{letter}1'].value: letter})
            temp_sheet[f"{letter}1"].value = sheet[f'{letter}1'].value
        else:
            break
        column_count += 1

    # Проходимся по массиву ИНН
    count = 0

    for code in codes:
        try:
            if len(code) > 10:
                print("ИПшки не принимаю")
                continue

            # Проверка на существование данного ИНН в базе
            is_exist = False
            index = 2
            while True:
                if not sheet[f"{headers['ИНН']}{index}"].value:
                    break
                elif sheet[f"{headers['ИНН']}{index}"].value == code:
                    is_exist = True
                    break
                index += 1

            if is_exist:
                continue
            else:
                count += 1

                current_data = parse.get_data(driver, wait, code)
                # Заполнение промежуточной таблицы
                if current_data:    
                    temp_sheet[f"{headers['№']}{count+1}"].value = count
                    temp_sheet[f"{headers['ИНН']}{count+1}"].value = code
                    temp_sheet[f"{headers['ОПФ']}{count+1}"].value = current_data["opf"]
                    temp_sheet[f"{headers['Организация']}{count+1}"].value = current_data["organization"]
                    temp_sheet[f"{headers['Контактное лицо']}{count+1}"].value = "-"
                    temp_sheet[f"{headers['E-mail потенциальный']}{count+1}"].value = current_data["email"]
                    temp_sheet[f"{headers['Директор']}{count+1}"].value = current_data["director"]
                    temp_sheet[f"{headers['Регион']}{count+1}"].value = current_data["address"]
                    temp_sheet[f"{headers['Страна']}{count+1}"].value = "Россия"

                    current_phones = current_data['phones']
                    phones_count = 1

                    if current_phones:
                        for i in range(len(current_phones)):
                            if current_phones[i]:
                                temp_sheet[f"{headers['Телефон{0}'.format(phones_count)]}{count+1}"].value = current_phones[i]
                                phones_count += 1
                            else:
                                break
                    else:
                        temp_sheet[f"{headers['Телефон1']}{index}"] = "-"

                    # Заполнение общей таблицы
                    sheet[f"{headers['№']}{index}"].value = index - 1
                    sheet[f"{headers['ИНН']}{index}"].value = code
                    sheet[f"{headers['ОПФ']}{index}"].value = current_data["opf"]
                    sheet[f"{headers['Организация']}{index}"].value = current_data["organization"]
                    sheet[f"{headers['Контактное лицо']}{index}"].value = "-"
                    sheet[f"{headers['E-mail потенциальный']}{index}"].value = current_data["email"]
                    sheet[f"{headers['Директор']}{index}"].value = current_data["director"]
                    sheet[f"{headers['Регион']}{index}"].value = current_data["address"]
                    sheet[f"{headers['Страна']}{index}"].value = "Россия"

                    phones_count = 1
                    if current_phones:
                        for i in range(len(current_phones)):
                            if current_phones[i]:
                                sheet[f"{headers['Телефон{0}'.format(phones_count)]}{index}"].value = current_phones[i]
                                phones_count += 1
                            else:
                                break
                    else:
                        sheet[f"{headers['Телефон1']}{index}"] = "-"
                else:
                    index -= 1
                    continue
        except Exception as e:
            index -= 1
            waste_codes.write(f'{code}\n')
            print(e)
            continue

    # Сохраняем значения в таблице
    waste_codes.close()

    temp_wb.save(dest_filename)
    workbook.save("database.xlsx")
    
    driver.quit()


